# coding: utf-8
import xlrd
import openpyxl as px
import re
class ExcelProcess:
    def __init__(self,inputFilePath,output_dir,prefix):
        #引数読み込み
        self.inputFilePath = inputFilePath
        self.output_dir = output_dir
        self.prefix = prefix
        #読み込み用にExcelファイルを読み込み
        f = xlrd.open_workbook(self.inputFilePath)
        self.sheet = f.sheets()
        #Headerを所得
        self.header = self.sheet[0].row_values(0)
        self.header = list(map(str, self.header))
        #書き込み用にExcelファイルを読み込み
        self.book = px.load_workbook(self.inputFilePath)
        self.pxSheet = self.book.worksheets[0]
        #proveanスコア,Predの行を所得
        self.proveanPredIndex, self.proveanPredCol = self.colAquire('PROVEAN_pred','both')
        self.proveanScoreIndex = self.colAquire('PROVEAN_score','index')
    def colAquire(self,target,out):
        #Headerにtargetがある場合はそのindexを所得、ない場合はHeaderの最後の列にtargetの行を追加
        try:
            index_num = self.header.index(target)
            #targetの行をリストで所得
            target_col = self.sheet[0].col(index_num)
            target_col = list(map(str,target_col))
        except:
            index_num = len(self.header)
            self.pxSheet.cell(column = index_num + 1, row = 1, value = target)
            self.header.append(target)
            target_col = [""] * self.colLenAquire()
            target_col[0] = target
        #outで指定したフォーマット(index or 行のリスト)でリターン
        if out == 'col':
            return target_col
        elif out == 'index':
            return index_num
        elif out == 'both':
            return index_num, target_col
    def colLenAquire(self):
        #Excelファイルの列数を所得
        infoCol = self.colAquire('INFO','col')
        colLen = len(infoCol)
        return colLen
    def writeExcel(self,result,colIndex):
        self.pxSheet.cell(column = self.proveanPredIndex + 1, row = colIndex + 1, value = result['resultPred'])
        self.pxSheet.cell(column = self.proveanScoreIndex + 1, row = colIndex + 1, value = result['resultValue'])
        self.book.save(self.output_dir + '/output_provean_' + self.prefix + ".xlsx")

